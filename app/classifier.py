import json
import base64
import logging
from openai import AsyncOpenAI

from app.config import settings

logger = logging.getLogger(__name__)

# Ollama Cloud exposes an OpenAI-compatible API at /v1
client = AsyncOpenAI(
    base_url=f"{settings.OLLAMA_BASE_URL}/v1",
    api_key=settings.OLLAMA_API_KEY,
)

CLASSIFICATION_PROMPT = """คุณเป็นผู้เชี่ยวชาญในการจำแนกเอกสาร วิเคราะห์เนื้อหาของเอกสารนี้แล้วจัดหมวดหมู่

หมวดหมู่ที่ใช้ได้:
{categories}

ตอบกลับเป็น JSON เท่านั้น ห้ามมีข้อความอื่น ตามรูปแบบนี้:
{{
    "category": "ชื่อหมวดหมู่",
    "subcategory": "หมวดย่อย (ถ้ามี)",
    "summary": "สรุปเนื้อหาสั้นๆ 1-2 ประโยค",
    "confidence": 0.95,
    "tags": ["tag1", "tag2", "tag3"],
    "key_info": {{
        "date": "วันที่ในเอกสาร (ถ้ามี)",
        "amount": "จำนวนเงิน (ถ้ามี)",
        "parties": ["ชื่อบุคคล/องค์กรที่เกี่ยวข้อง"]
    }}
}}
"""


async def classify_text_document(text_content: str, filename: str) -> dict:
    """Classify a text-based document using Ollama."""
    categories_str = "\n".join(f"- {c}" for c in settings.CATEGORIES)

    try:
        response = await client.chat.completions.create(
            model=settings.OLLAMA_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": "คุณเป็น AI ที่จำแนกเอกสารเป็นหมวดหมู่ ตอบเป็น JSON เท่านั้น ห้ามมีข้อความอื่นนอกจาก JSON",
                },
                {
                    "role": "user",
                    "content": (
                        f"{CLASSIFICATION_PROMPT.format(categories=categories_str)}\n\n"
                        f"ชื่อไฟล์: {filename}\n\n"
                        f"เนื้อหาเอกสาร:\n{text_content[:4000]}"
                    ),
                },
            ],
            temperature=0.1,
        )

        result_text = response.choices[0].message.content
        # Extract JSON from response
        if "```json" in result_text:
            result_text = result_text.split("```json")[1].split("```")[0]
        elif "```" in result_text:
            result_text = result_text.split("```")[1].split("```")[0]

        return json.loads(result_text.strip())

    except Exception as e:
        logger.error(f"Classification error: {e}")
        return {
            "category": "อื่นๆ",
            "subcategory": "",
            "summary": f"ไม่สามารถจำแนกได้: {filename}",
            "confidence": 0.0,
            "tags": [],
            "key_info": {},
        }


async def classify_image(image_bytes: bytes, filename: str, mime_type: str = "image/jpeg") -> dict:
    """Classify an image document using Ollama vision model."""
    categories_str = "\n".join(f"- {c}" for c in settings.CATEGORIES)
    b64_image = base64.b64encode(image_bytes).decode("utf-8")

    media_type = mime_type if mime_type in ["image/jpeg", "image/png", "image/gif", "image/webp"] else "image/jpeg"

    try:
        response = await client.chat.completions.create(
            model=settings.OLLAMA_VISION_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": "คุณเป็น AI ที่จำแนกเอกสารเป็นหมวดหมู่ ตอบเป็น JSON เท่านั้น ห้ามมีข้อความอื่นนอกจาก JSON",
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{media_type};base64,{b64_image}",
                            },
                        },
                        {
                            "type": "text",
                            "text": (
                                f"{CLASSIFICATION_PROMPT.format(categories=categories_str)}\n\n"
                                f"ชื่อไฟล์: {filename}\n"
                                "วิเคราะห์รูปภาพนี้และจำแนกหมวดหมู่"
                            ),
                        },
                    ],
                },
            ],
            temperature=0.1,
        )

        result_text = response.choices[0].message.content
        if "```json" in result_text:
            result_text = result_text.split("```json")[1].split("```")[0]
        elif "```" in result_text:
            result_text = result_text.split("```")[1].split("```")[0]

        return json.loads(result_text.strip())

    except Exception as e:
        logger.error(f"Image classification error: {e}")
        return {
            "category": "รูปภาพ/หลักฐาน",
            "subcategory": "",
            "summary": f"ไม่สามารถวิเคราะห์รูปภาพได้: {filename}",
            "confidence": 0.0,
            "tags": [],
            "key_info": {},
        }


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text content from PDF."""
    try:
        from PyPDF2 import PdfReader
        import io
        reader = PdfReader(io.BytesIO(file_bytes))
        text_parts = []
        for page in reader.pages[:10]:  # Limit to 10 pages
            text_parts.append(page.extract_text() or "")
        return "\n".join(text_parts)
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
        return ""


def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extract text content from DOCX."""
    try:
        from docx import Document
        import io
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        logger.error(f"DOCX extraction error: {e}")
        return ""


def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extract text content from XLSX."""
    try:
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        text_parts = []
        for sheet in wb.sheetnames[:3]:  # Limit to 3 sheets
            ws = wb[sheet]
            text_parts.append(f"[Sheet: {sheet}]")
            for row in ws.iter_rows(max_row=50, values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        logger.error(f"XLSX extraction error: {e}")
        return ""


async def classify_document(file_bytes: bytes, filename: str, file_type: str, mime_type: str) -> dict:
    """Main classification entry point — routes to appropriate classifier."""
    if file_type in ("jpg", "jpeg", "png", "gif", "webp", "image"):
        return await classify_image(file_bytes, filename, mime_type)

    # Extract text based on file type
    text = ""
    if file_type == "pdf":
        text = extract_text_from_pdf(file_bytes)
        if not text.strip():
            # PDF might be image-based — use vision
            return await classify_image(file_bytes, filename, "image/png")
    elif file_type in ("docx", "doc"):
        text = extract_text_from_docx(file_bytes)
    elif file_type in ("xlsx", "xls"):
        text = extract_text_from_xlsx(file_bytes)
    else:
        text = file_bytes.decode("utf-8", errors="ignore")[:4000]

    if text.strip():
        return await classify_text_document(text, filename)

    return {
        "category": "อื่นๆ",
        "subcategory": "",
        "summary": f"ไม่สามารถอ่านเนื้อหาได้: {filename}",
        "confidence": 0.0,
        "tags": [],
        "key_info": {},
    }
